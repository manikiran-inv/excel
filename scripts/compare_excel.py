from openpyxl import load_workbook
import sys

old_file = sys.argv[1]
new_file = sys.argv[2]
output_file = sys.argv[3]

wb_old = load_workbook(old_file, data_only=False)
wb_new = load_workbook(new_file, data_only=False)

changes = []

for sheet in wb_old.sheetnames:
    if sheet not in wb_new.sheetnames:
        continue

    ws_old = wb_old[sheet]
    ws_new = wb_new[sheet]

    max_row = max(ws_old.max_row, ws_new.max_row)
    max_col = max(ws_old.max_column, ws_new.max_column)

    for r in range(1, max_row + 1):
        for c in range(1, max_col + 1):

            old = ws_old.cell(r, c).value
            new = ws_new.cell(r, c).value

            if old != new:
                changes.append([
                    sheet,
                    ws_old.cell(r, c).coordinate,
                    old,
                    new
                ])

with open(output_file, "w", encoding="utf-8") as f:

    f.write("# Excel Difference Report\n\n")

    if not changes:
        f.write("✅ No differences found.\n")
    else:
        f.write("| Sheet | Cell | Old | New |\n")
        f.write("|------|------|------|------|\n")

        for sheet, cell, old, new in changes:
            f.write(
                f"| {sheet} | {cell} | {old} | {new} |\n"
            )

print(f"Found {len(changes)} changes.")